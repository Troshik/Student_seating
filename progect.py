import sys
import sqlite3
import openpyxl
import random
import xlwt


from PyQt5 import QtCore, QtWidgets
from PyQt5.QtGui import QColor
from PyQt5.QtWidgets import QMainWindow, QTableWidgetItem, QWidget, QApplication, QDialog, QMessageBox, QFileDialog


class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(590, 360)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.gridLayoutWidget = QtWidgets.QWidget(self.centralwidget)
        self.gridLayoutWidget.setGeometry(QtCore.QRect(20, 20, 551, 301))
        self.gridLayoutWidget.setObjectName("gridLayoutWidget")
        self.gridLayout = QtWidgets.QGridLayout(self.gridLayoutWidget)
        self.gridLayout.setContentsMargins(0, 0, 0, 0)
        self.gridLayout.setObjectName("gridLayout")
        self.gridLayout_2 = QtWidgets.QGridLayout()
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.button_stud_data = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.button_stud_data.setObjectName("button_stud_data")
        self.gridLayout_2.addWidget(self.button_stud_data, 0, 0, 1, 1)
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        self.button_yes_presence_data = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.button_yes_presence_data.setObjectName("button_yes_presence_data")
        self.horizontalLayout_2.addWidget(self.button_yes_presence_data)
        self.button_no_presence_data = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.button_no_presence_data.setObjectName("button_no_presence_data")
        self.horizontalLayout_2.addWidget(self.button_no_presence_data)
        self.gridLayout_2.addLayout(self.horizontalLayout_2, 1, 0, 1, 1)
        self.gridLayout.addLayout(self.gridLayout_2, 2, 0, 1, 1)
        self.gridLayout_3 = QtWidgets.QGridLayout()
        self.gridLayout_3.setObjectName("gridLayout_3")
        self.textEdit_add_stud = QtWidgets.QTextEdit(self.gridLayoutWidget)
        self.textEdit_add_stud.setObjectName("textEdit_add_stud")
        self.gridLayout_3.addWidget(self.textEdit_add_stud, 1, 0, 1, 1)
        self.button_add_stud = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.button_add_stud.setObjectName("button_add_stud")
        self.gridLayout_3.addWidget(self.button_add_stud, 1, 1, 1, 1)
        self.label_format = QtWidgets.QLabel(self.gridLayoutWidget)
        self.label_format.setObjectName("label_format")
        self.gridLayout_3.addWidget(self.label_format, 0, 0, 1, 1)
        self.gridLayout.addLayout(self.gridLayout_3, 4, 0, 1, 1)
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.label_data_aud = QtWidgets.QLabel(self.gridLayoutWidget)
        self.label_data_aud.setObjectName("label_data_aud")
        self.horizontalLayout.addWidget(self.label_data_aud)
        self.lineEdit_data_aud = QtWidgets.QLineEdit(self.gridLayoutWidget)
        self.lineEdit_data_aud.setObjectName("lineEdit_data_aud")
        self.horizontalLayout.addWidget(self.lineEdit_data_aud)
        self.button_data_aud = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.button_data_aud.setObjectName("button_data_aud")
        self.horizontalLayout.addWidget(self.button_data_aud)
        self.gridLayout.addLayout(self.horizontalLayout, 0, 0, 1, 1)
        self.button_save = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.button_save.setObjectName("button_save")
        self.gridLayout.addWidget(self.button_save, 5, 0, 1, 1)
        self.horizontalLayout_3 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_3.setObjectName("horizontalLayout_3")
        self.label_stud = QtWidgets.QLabel(self.gridLayoutWidget)
        self.label_stud.setObjectName("label_stud")
        self.horizontalLayout_3.addWidget(self.label_stud)
        self.lineEdit_stud = QtWidgets.QLineEdit(self.gridLayoutWidget)
        self.lineEdit_stud.setObjectName("lineEdit_stud")
        self.horizontalLayout_3.addWidget(self.lineEdit_stud)
        self.button_stud = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.button_stud.setObjectName("button_stud")
        self.horizontalLayout_3.addWidget(self.button_stud)
        self.gridLayout.addLayout(self.horizontalLayout_3, 1, 0, 1, 1)
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 588, 21))
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.button_stud_data.setText(_translate("MainWindow", "Список учеников"))
        self.button_yes_presence_data.setText(_translate("MainWindow", "Присутствующие ученики"))
        self.button_no_presence_data.setText(_translate("MainWindow", "Отсутствующие ученики"))
        self.button_add_stud.setText(_translate("MainWindow", "Добавить уечника"))
        self.label_format.setText(
            _translate("MainWindow", "Формат ввода: \'Фамилия:Имя:Отчество:Класс:Школа\' (без кавычек)"))
        self.label_data_aud.setText(_translate("MainWindow", "  Рассадка в аудитории N: "))
        self.button_data_aud.setText(_translate("MainWindow", "Открыть"))
        self.button_save.setText(_translate("MainWindow", "Сохранить"))
        self.label_stud.setText(_translate("MainWindow", "   Поиск ученика:                "))
        self.button_stud.setText(_translate("MainWindow", "Поиск"))


class Ui_Dialog(object):
    def setupUi(self, Dialog):
        Dialog.setObjectName("Dialoq")
        Dialog.resize(700, 640)

        self.pushButton_change = QtWidgets.QPushButton(Dialog)
        self.pushButton_change.setGeometry(QtCore.QRect(45, 540, 600, 32))
        self.pushButton_change.setObjectName("pushButton_change")

        self.tableWidget = QtWidgets.QTableWidget(Dialog)
        self.tableWidget.setGeometry(QtCore.QRect(30, 30, 650, 500))
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setRowCount(0)
        self.tableWidget.setColumnCount(9)
        self.tableWidget.verticalHeader().setVisible(False)
        self.tableWidget.setHorizontalHeaderLabels(["Номер", "Фамилия", "Имя", "Отчество", "Класс", "Школа",
                                                    "Аудитория", "Место", "Присутствие"])

        self.retranslateUi(Dialog)
        QtCore.QMetaObject.connectSlotsByName(Dialog)

    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Dialog", "Dialog"))
        self.pushButton_change.setText(_translate("MainWindow", "Изменить"))


class Data_presence(QDialog, Ui_Dialog):

    def __init__(self):
        super().__init__()
        self.setupUi(self)


class Ui_Form(object):
    def setupUi(self, form):
        form.setObjectName("Form")
        form.resize(549, 475)

        self.tableWidget = QtWidgets.QTableWidget(form)
        self.tableWidget.setGeometry(QtCore.QRect(10, 10, 521, 441))
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setRowCount(0)
        self.tableWidget.setColumnCount(9)
        self.tableWidget.verticalHeader().setVisible(False)
        self.tableWidget.setHorizontalHeaderLabels(["Номер", "Фамилия", "Имя", "Отчество", "Класс", "Школа",
                                                    "Аудитория", "Место", "Присутствие"])

        self.retranslateUi(form)
        QtCore.QMetaObject.connectSlotsByName(form)

    def retranslateUi(self, form):
        _translate = QtCore.QCoreApplication.translate
        form.setWindowTitle(_translate("Form", "Form"))


class Table_presence(QWidget, Ui_Form):

    def __init__(self):
        super().__init__()
        self.setupUi(self)


class MyWidget(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.w = Data_presence()
        self.t = Table_presence()

        self.setupUi(self)
        a = False
        while not a:
            try:
                self.open_file_dialog()
                self.audiences = {}
                self.con = sqlite3.connect(':memory:')
                self.export_to_sqlite()
                self.add_index()
                self.add_audienc_place()
                a = True
            except:
                a = False
                QMessageBox.about(self, "Ошибка", "Введены некорректные данные")

        self.button_stud_data.clicked.connect(self.open_data)
        self.w.pushButton_change.clicked.connect(self.update_elems)
        self.button_data_aud.clicked.connect(self.open_data_aud)
        self.button_no_presence_data.clicked.connect(self.no_presence_data)
        self.button_yes_presence_data.clicked.connect(self.yes_presence_data)
        self.button_add_stud.clicked.connect(self.add_stud)
        self.button_stud.clicked.connect(self.search)
        self.button_save.clicked.connect(self.save_xlsx)
        self.setWindowTitle('Рассадка учеников')

    def open_file_dialog(self):
        self.stud_filename, ok = QFileDialog.getOpenFileName(
            self,
            "Выберите файл со списком учеников",
            "*.xlsx"
        )
        self.aud_filename, ok = QFileDialog.getOpenFileName(
            self,
            "Выберите файл со списком аудиторий",
            "*.xlsx"
        )

    def export_to_sqlite(self):
        cursor = self.con.cursor()
        cursor.execute('CREATE TABLE IF NOT EXISTS Список_учеников'
                       ' (id inteqer, surname text, name text, patronymic text, class text, school text,'
                       ' audienc text, place text, presence text)')
        file_to_read = openpyxl.load_workbook(self.stud_filename, data_only=True)
        sheet = file_to_read['Лист1']
        for row in range(1, sheet.max_row + 1):
            data = []
            for col in range(1, 8):
                value = sheet.cell(row, col).value
                data.append(value)
            cursor.execute("INSERT INTO Список_учеников VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                           (None, data[0], data[1], data[2], data[3], data[4], data[5], data[6], 'No'))
            self.con.commit()

    def creating_audiences(self):
        wb = openpyxl.load_workbook(self.aud_filename)
        ws = wb['Лист1']
        column = ws['A']
        column_a = [column[x].value for x in range(len(column))]
        column = ws['B']
        column_b = [column[x].value for x in range(len(column))]
        for i in range(1, len(column_a)):
            self.audiences[str(column_a[i])] = str(column_b[i])

    def add_index(self):
        cursor = self.con.cursor()
        cursor.execute("""SELECT * FROM Список_учеников ORDER BY school, class, surname""")

        rows = cursor.fetchall()
        for i in range(len(tuple(rows))):
            upd = """UPDATE Список_учеников SET id = ? where surname = ? and name = ? and school = ?"""
            data = (i + 1, rows[i][1], rows[i][2], rows[i][5])
            cursor.execute(upd, data)
        self.con.commit()

    def add_audienc_place(self):
        self.creating_audiences()
        cursor = self.con.cursor()
        cursor.execute("""SELECT * FROM Список_учеников ORDER BY id""")
        name_aud = list(self.audiences)

        # заполняем колонну audienc
        a = 0
        b = []
        for elem in self.audiences.items():
            b.append(int(elem[1]) * 3)
        rows = cursor.fetchall()
        for i in range(len(tuple(rows))):
            upd = """UPDATE Список_учеников SET audienc = ? where id = ?"""
            data = (name_aud[a], i + 1)
            cursor.execute(upd, data)
            b[a] = b[a] - 1
            if b[a] == 0:
                del(b[a])
                del(name_aud[a])
                a -= 1
            if a < len(b) - 1:
                a += 1
            else:
                a = 0
        self.con.commit()

        name_aud = list(self.audiences)

        # заполняем колонну place
        for aud in name_aud:
            aud_data = cursor.execute("""SELECT * FROM Список_учеников WHERE audienc = ?""", (str(aud),)).fetchall()
            aud_list = [[], [], []]
            for i in range(int(self.audiences[aud])):
                aud_list[0].append('A' + str(i + 1))
                aud_list[1].append('B' + str(i + 1))
                aud_list[2].append('C' + str(i + 1))
            for i in range(len(tuple(aud_data))):
                a = random.randint(1, len(aud_list))
                b = random.randint(1, len(aud_list[a - 1]))
                p_id = aud_data[i][0]
                place = aud_list[a - 1][b - 1]
                upd = """UPDATE Список_учеников SET place = ? where id = ?"""
                data = (place, p_id)
                cursor.execute(upd, data)
                self.con.commit()
                del(aud_list[a - 1][b - 1])
                if len(aud_list[a - 1]) == 0:
                    del aud_list[a - 1]

    def open_data(self):
        self.w.setWindowTitle('Список учеников')
        cur = self.con.cursor()
        tab = cur.execute("""SELECT * FROM Список_учеников ORDER BY id""").fetchall()
        a = 0
        for elem in tab:
            a += 1
        self.w.tableWidget.setRowCount(a)
        for i, row in enumerate(tab):
            for j, elem in enumerate(row):
                self.w.tableWidget.setItem(
                    i, j, QTableWidgetItem(str(elem)))

        self.w.show()
        self.statusBar().showMessage("Список учеников")

    def generate_new_elems(self, ids):
        cur = self.con.cursor()
        for i in ids:
            tab = cur.execute("""SELECT * FROM Список_учеников ORDER BY id""")
            rows = []
            for row in tab:
                rows.append(row)
            if rows[int(i) - 1][8] == 'No':
                presenc = ('Yes', i)
            else:
                presenc = ('No', i)
            upd = """Update Список_учеников set presence = ? where id = ?"""
            cur.execute(upd, presenc)
            self.con.commit()

    def update_elems(self):
        rows = list(set([i.row() for i in self.w.tableWidget.selectedItems()]))
        ids = [self.w.tableWidget.item(i, 0).text() for i in rows]
        valid = QMessageBox.question(self, '', "Действительно заменить элементы с id " + ",".join(ids), QMessageBox.Yes,
                                     QMessageBox.No)
        if valid == QMessageBox.Yes:
            self.generate_new_elems(ids)
        self.w.close()

    def color_row(self, row, color):
        for i in range(self.t.tableWidget.columnCount()):
            self.t.tableWidget.item(row, i).setBackground(color)

    def yes_presence_data(self):
        cur = self.con.cursor()
        tab = cur.execute("""SELECT * FROM Список_учеников WHERE presence = ? ORDER BY id""", ('Yes',)).fetchall()
        a = 0
        for elem in tab:
            a += 1
        self.t.tableWidget.setColumnCount(8)
        self.t.tableWidget.setRowCount(a)
        for i, row in enumerate(tab):
            for j, elem in enumerate(row):
                self.t.tableWidget.setItem(
                    i, j, QTableWidgetItem(str(elem)))

        self.t.setWindowTitle('Список присутствующих учеников')
        self.t.show()
        self.statusBar().showMessage("Список присутствующих учеников")

    def no_presence_data(self):
        cur = self.con.cursor()
        tab = cur.execute("""SELECT * FROM Список_учеников WHERE presence = ? ORDER BY id""", ('No',)).fetchall()
        a = 0
        for elem in tab:
            a += 1
        self.t.tableWidget.setColumnCount(8)
        self.t.tableWidget.setRowCount(a)
        for i, row in enumerate(tab):
            for j, elem in enumerate(row):
                self.t.tableWidget.setItem(
                    i, j, QTableWidgetItem(str(elem)))

        self.t.setWindowTitle('Список отсутствующих учеников')
        self.t.show()
        self.statusBar().showMessage("Список отсутствующих учеников")

    def search(self):
        cur = self.con.cursor()
        item_stud = list(self.lineEdit_stud.text().split())
        try:
            result = cur.execute("SELECT * FROM Список_учеников WHERE surname = ? and name = ?  ORDER BY id",
                                 item_stud).fetchall()

            self.t.tableWidget.setRowCount(len(result))
            for i, elem in enumerate(result):
                for j, val in enumerate(elem):
                    self.t.tableWidget.setItem(i, j, QTableWidgetItem(str(val)))

            self.t.setWindowTitle(f"Поиск ученика {' '.join(item_stud)}.")
            self.t.show()
            self.statusBar().showMessage(f"Поиск ученика {' '.join(item_stud)}.")

        except:
            self.statusBar().showMessage("Ошибка: Вам нужно указать имя и фамилию ученика")

    def open_data_aud(self):
        cur = self.con.cursor()
        item_aud = str(self.lineEdit_data_aud.text())
        if item_aud in list(self.audiences):

            result = cur.execute("SELECT * FROM Список_учеников WHERE audienc = ?  ORDER BY id", (item_aud,)).fetchall()

            self.t.tableWidget.setRowCount(len(result))
            for i, elem in enumerate(result):
                for j, val in enumerate(elem):
                    self.t.tableWidget.setItem(i, j, QTableWidgetItem(str(val)))
                if elem[8] == "No":
                    self.color_row(i, QColor(220, 40, 40))  # отсутствующий ученик помечается красным цветом

            self.t.setWindowTitle(f'Список учеников в аудитории {item_aud}')
            self.t.show()
            self.statusBar().showMessage(f"Список учеников в аудитории {item_aud}."
                                         f" Красным цветом помечены отсутствующие ученики")

        else:
            self.statusBar().showMessage("Ошибка: такой аудитории не существует")

    def add_stud(self):
        cur = self.con.cursor()
        query = self.lineEdit_stud.setObjectName
        try:
            data = query.split(':')
            cur.execute("INSERT INTO Список_учеников VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        (None, data[0], data[1], data[2], data[3], data[4], None, None, 'No'))
            self.add_index()
            self.add_audienc_place()
            self.open_data()
        except:
            self.statusBar().showMessage("Ошибка: введены данные неверного формата")

    def save_xlsx(self):
        cur = self.con.cursor()
        rows = cur.execute("SELECT * FROM Список_учеников ORDER BY id").fetchall()

        book = xlwt.Workbook(encoding="utf-8")
        sheet1 = book.add_sheet("Лист1")

        row = 0
        col = 0

        for row_data in tuple(rows):
            sheet1.write(row, col, row_data[1])
            sheet1.write(row, col + 1, row_data[2])
            sheet1.write(row, col + 2, row_data[3])
            sheet1.write(row, col + 3, row_data[4])
            sheet1.write(row, col + 4, row_data[5])
            sheet1.write(row, col + 5, row_data[6])
            sheet1.write(row, col + 6, row_data[7])
            row += 1

        name, ok = QFileDialog.getSaveFileName(
            self,
            "Куда сохранить?", "Рассадка_учеников.xls"
        )
        book.save(name)
        self.statusBar().showMessage("Таблица сохранена")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MyWidget()
    ex.show()
    sys.exit(app.exec())
